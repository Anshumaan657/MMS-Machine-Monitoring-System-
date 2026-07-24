from collections import Counter, defaultdict
from datetime import datetime
import json
from pathlib import Path
import re
import sys

from openpyxl import load_workbook


NULL_MARKERS = {"", "NULL", "NONE", "N/A", "NA", "-"}


def text(value):
    return "" if value is None else str(value).strip()


def is_missing(value):
    return text(value).upper() in NULL_MARKERS


def number(value):
    raw = text(value).replace(",", "")
    if not raw or raw.upper() in NULL_MARKERS:
        return None
    try:
        return float(raw)
    except ValueError:
        return None


def timestamp(value):
    raw = text(value).replace("  ", " ")
    for pattern in (
        "%d/%m/%Y %I:%M:%S %p",
        "%d/%m/%Y %I:%M %p",
        "%d/%m/%Y",
        "%Y-%m-%d %H:%M:%S",
    ):
        try:
            return datetime.strptime(raw, pattern)
        except ValueError:
            continue
    return None


def hours(value):
    match = re.fullmatch(r"(\d+):(\d{1,2})(?::(\d{1,2}))?", text(value))
    if not match:
        return None
    hour, minute, second = match.groups()
    return int(hour) + int(minute) / 60 + int(second or 0) / 3600


def rows_from(sheet, header_row=6):
    headers = [text(cell.value) for cell in next(sheet.iter_rows(min_row=header_row, max_row=header_row))]
    rows = []
    for values in sheet.iter_rows(min_row=header_row + 1, values_only=True):
        row = dict(zip(headers, values))
        if any(not is_missing(value) for value in values):
            rows.append(row)
    return headers, rows


source = Path(sys.argv[1])
destination = Path(sys.argv[2])
book = load_workbook(source, read_only=True, data_only=True)

_, product_rows = rows_from(book["Product Log Book"])
product_rows = [row for row in product_rows if text(row["Part No."]).upper() != "TOTAL === >"]

_, downtime_rows = rows_from(book["Down Time Details"])
downtime_rows = [row for row in downtime_rows if text(row["Shift"]).upper() != "TOTAL"]

machine_stats = defaultdict(
    lambda: {
        "production": 0.0,
        "target": 0.0,
        "productRecords": 0,
        "downtimeHours": 0.0,
        "revenueLoss": 0.0,
        "downtimeEvents": 0,
        "unreportedEvents": 0,
    }
)
shift_stats = defaultdict(
    lambda: {
        "production": 0.0,
        "target": 0.0,
        "downtimeHours": 0.0,
        "revenueLoss": 0.0,
    }
)
monthly_stats = defaultdict(
    lambda: {
        "production": 0.0,
        "target": 0.0,
        "downtimeHours": 0.0,
        "revenueLoss": 0.0,
    }
)
daily_stats = defaultdict(
    lambda: {
        "production": 0.0,
        "target": 0.0,
        "downtimeHours": 0.0,
        "revenueLoss": 0.0,
    }
)

product_dates = []
for row in product_rows:
    machine = text(row["Machine"])
    shift = text(row["Shift"])
    date_value = timestamp(row["Date"])
    qty = number(row["Qty"]) or 0
    target = number(row["Shift Target"]) or 0

    if date_value:
        product_dates.append(date_value)
        month = date_value.strftime("%Y-%m")
        day = date_value.strftime("%Y-%m-%d")
        monthly_stats[month]["production"] += qty
        monthly_stats[month]["target"] += target
        daily_stats[day]["production"] += qty
        daily_stats[day]["target"] += target

    machine_stats[machine]["production"] += qty
    machine_stats[machine]["target"] += target
    machine_stats[machine]["productRecords"] += 1
    shift_stats[shift]["production"] += qty
    shift_stats[shift]["target"] += target

downtime_dates = []
invalid_durations = 0
for row in downtime_rows:
    machine = text(row["Machine"])
    shift = text(row["Shift"])
    date_value = timestamp(row["Date"])
    duration = hours(row["Duration"])
    loss = number(row["Revenue"]) or 0
    reason = text(row["Reason"]).upper()

    if duration is None:
        invalid_durations += 1
        duration = 0
    if date_value:
        downtime_dates.append(date_value)
        month = date_value.strftime("%Y-%m")
        day = date_value.strftime("%Y-%m-%d")
        monthly_stats[month]["downtimeHours"] += duration
        monthly_stats[month]["revenueLoss"] += loss
        daily_stats[day]["downtimeHours"] += duration
        daily_stats[day]["revenueLoss"] += loss

    machine_stats[machine]["downtimeHours"] += duration
    machine_stats[machine]["revenueLoss"] += loss
    machine_stats[machine]["downtimeEvents"] += 1
    if reason == "UNREPORTED":
        machine_stats[machine]["unreportedEvents"] += 1
    shift_stats[shift]["downtimeHours"] += duration
    shift_stats[shift]["revenueLoss"] += loss


def rounded(value, digits=1):
    return round(value, digits)


machines = []
for machine, values in machine_stats.items():
    if not machine:
        continue
    target = values["target"]
    events = values["downtimeEvents"]
    machines.append(
        {
            "machine": machine,
            "production": int(round(values["production"])),
            "target": rounded(target),
            "attainment": rounded(values["production"] / target * 100) if target else None,
            "downtimeHours": rounded(values["downtimeHours"]),
            "revenueLoss": int(round(values["revenueLoss"])),
            "downtimeEvents": events,
            "unreportedRate": rounded(values["unreportedEvents"] / events * 100, 2) if events else 0,
        }
    )

machines.sort(key=lambda item: item["downtimeHours"], reverse=True)

shifts = []
for shift, values in sorted(shift_stats.items()):
    if not shift:
        continue
    target = values["target"]
    shifts.append(
        {
            "shift": shift,
            "production": int(round(values["production"])),
            "target": rounded(target),
            "attainment": rounded(values["production"] / target * 100) if target else None,
            "downtimeHours": rounded(values["downtimeHours"]),
            "revenueLoss": int(round(values["revenueLoss"])),
        }
    )

months = []
for month, values in sorted(monthly_stats.items()):
    target = values["target"]
    months.append(
        {
            "month": month,
            "production": int(round(values["production"])),
            "target": rounded(target),
            "attainment": rounded(values["production"] / target * 100) if target else None,
            "downtimeHours": rounded(values["downtimeHours"]),
            "revenueLoss": int(round(values["revenueLoss"])),
        }
    )

latest_day = max(daily_stats)
latest = daily_stats[latest_day]
latest_target = latest["target"]
latest_machine_rows = defaultdict(
    lambda: {"production": 0.0, "target": 0.0, "downtimeHours": 0.0, "revenueLoss": 0.0}
)
for row in product_rows:
    date_value = timestamp(row["Date"])
    if date_value and date_value.strftime("%Y-%m-%d") == latest_day:
        machine = text(row["Machine"])
        latest_machine_rows[machine]["production"] += number(row["Qty"]) or 0
        latest_machine_rows[machine]["target"] += number(row["Shift Target"]) or 0
for row in downtime_rows:
    date_value = timestamp(row["Date"])
    if date_value and date_value.strftime("%Y-%m-%d") == latest_day:
        machine = text(row["Machine"])
        latest_machine_rows[machine]["downtimeHours"] += hours(row["Duration"]) or 0
        latest_machine_rows[machine]["revenueLoss"] += number(row["Revenue"]) or 0
latest_machine = max(latest_machine_rows.items(), key=lambda item: item[1]["downtimeHours"])

unreported_count = sum(text(row["Reason"]).upper() == "UNREPORTED" for row in downtime_rows)
no_operator_product = sum("NO OPERATOR" in text(row["Operator"]).upper() for row in product_rows)
no_operator_downtime = sum("NO OPERATOR" in text(row["Operator Name"]).upper() for row in downtime_rows)
missing_products = sum(is_missing(row["Product Name"]) for row in product_rows)
missing_downtime_products = sum(is_missing(row["Product Name"]) for row in downtime_rows)
zero_reject = sum((number(row["Reject Qty"]) or 0) == 0 for row in product_rows)
zero_rework = sum((number(row["Rework Qty"]) or 0) == 0 for row in product_rows)
total_downtime = sum(item["downtimeHours"] for item in machines)
total_loss = sum(item["revenueLoss"] for item in machines)
total_production = sum(item["production"] for item in machines)
total_target = sum(item["target"] for item in machines)

summary = {
    "source": {
        "company": text(book["Product Log Book"]["A1"].value),
        "fileName": source.stem.replace(".xlsx", "") + ".xls",
        "generatedAt": datetime.now().isoformat(timespec="seconds"),
        "productDateRange": [
            min(product_dates).strftime("%Y-%m-%d"),
            max(product_dates).strftime("%Y-%m-%d"),
        ],
        "downtimeDateRange": [
            min(downtime_dates).strftime("%Y-%m-%d"),
            max(downtime_dates).strftime("%Y-%m-%d"),
        ],
    },
    "overview": {
        "machines": len(machines),
        "productRecords": len(product_rows),
        "downtimeEvents": len(downtime_rows),
        "totalProduction": total_production,
        "totalTarget": rounded(total_target),
        "targetAttainment": rounded(total_production / total_target * 100) if total_target else None,
        "downtimeHours": rounded(total_downtime),
        "reportedRevenueLoss": total_loss,
    },
    "quality": {
        "unreportedDowntimeEvents": unreported_count,
        "unreportedDowntimeRate": rounded(unreported_count / len(downtime_rows) * 100, 2),
        "missingProductRecords": missing_products,
        "missingDowntimeProducts": missing_downtime_products,
        "noOperatorProductRecords": no_operator_product,
        "noOperatorDowntimeEvents": no_operator_downtime,
        "invalidDurations": invalid_durations,
        "zeroRejectRecords": zero_reject,
        "zeroReworkRecords": zero_rework,
    },
    "machines": machines,
    "shifts": shifts,
    "monthly": months,
    "latestDay": {
        "date": latest_day,
        "production": int(round(latest["production"])),
        "target": rounded(latest_target),
        "attainment": rounded(latest["production"] / latest_target * 100) if latest_target else None,
        "downtimeHours": rounded(latest["downtimeHours"]),
        "reportedRevenueLoss": int(round(latest["revenueLoss"])),
        "topDowntimeMachine": latest_machine[0],
        "topDowntimeMachineHours": rounded(latest_machine[1]["downtimeHours"]),
    },
}

destination.parent.mkdir(parents=True, exist_ok=True)
destination.write_text(json.dumps(summary, indent=2), encoding="utf-8")
print(destination)
